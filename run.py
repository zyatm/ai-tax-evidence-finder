#!/usr/bin/env python3
"""
10-K Evidence Extraction CLI

Usage:
    python run.py extract <pdf_file> [--output <output_dir>]
    python run.py batch <pdf_dir> [--output <output_dir>]

Examples:
    python run.py extract company_10K.pdf
    python run.py extract company_10K.pdf --output ./results
    python run.py batch ./10k_pdfs --output ./results
"""

import sys
import os
import argparse
from pathlib import Path

# Add src to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

from stage1_parser import parse_document
from stage2_verbatim import VerbatimExtractor


def extract_single(pdf_path: str, output_dir: str = None, config_path: str = None):
    """Extract evidence from a single 10-K PDF."""
    pdf_path = Path(pdf_path)

    if not pdf_path.exists():
        print(f"Error: File not found: {pdf_path}")
        sys.exit(1)

    output_dir = Path(output_dir) if output_dir else pdf_path.parent
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*60}")
    print(f"Extracting: {pdf_path.name}")
    print(f"Output: {output_dir}")
    if config_path:
        print(f"Config: {config_path}")
    print(f"{'='*60}\n")

    # Parse
    doc = parse_document(str(pdf_path))

    # Extract
    extractor = VerbatimExtractor(config_path=config_path)
    result = extractor.extract(doc, pdf_path.stem)
    
    # Save JSON
    import json
    json_path = output_dir / f"{pdf_path.stem}_extraction.json"
    with open(json_path, 'w') as f:
        json.dump(result.to_dict(), f, indent=2)
    
    # Generate Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Evidence Binder"
    
    headers = ["Block", "Category", "Evidence Text", "Page", "Section", "Keyword", "Confidence", "Verified"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    row = 2
    for extraction in result.extractions:
        if not extraction.evidence:
            ws.cell(row=row, column=1, value=extraction.block)
            ws.cell(row=row, column=2, value=extraction.category)
            ws.cell(row=row, column=3, value="(no evidence found)")
            row += 1
        else:
            for ev in extraction.evidence:
                ws.cell(row=row, column=1, value=extraction.block)
                ws.cell(row=row, column=2, value=extraction.category)
                ws.cell(row=row, column=3, value=ev.text)
                ws.cell(row=row, column=4, value=ev.page)
                ws.cell(row=row, column=5, value=ev.section)
                ws.cell(row=row, column=6, value=ev.match_keyword)
                ws.cell(row=row, column=7, value=ev.confidence)
                ws.cell(row=row, column=8, value="✓" if ev.verified else "○")
                row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 80
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    
    for row_num in range(2, row):
        ws.cell(row=row_num, column=3).alignment = Alignment(wrap_text=True, vertical='top')
    
    xlsx_path = output_dir / f"{pdf_path.stem}_evidence_binder.xlsx"
    wb.save(xlsx_path)
    
    print(f"\n{'='*60}")
    print(f"COMPLETE")
    print(f"{'='*60}")
    print(f"Evidence items: {result.total_evidence}")
    print(f"Verified: {result.verified_count}/{result.total_evidence}")
    print(f"Cost: ${result.cost_estimate:.4f}")
    print(f"\nOutputs:")
    print(f"  JSON: {json_path}")
    print(f"  Excel: {xlsx_path}")
    
    return result


def extract_batch(pdf_dir: str, output_dir: str = None, config_path: str = None):
    """Extract evidence from all PDFs in a directory."""
    pdf_dir = Path(pdf_dir)

    if not pdf_dir.exists():
        print(f"Error: Directory not found: {pdf_dir}")
        sys.exit(1)

    pdf_files = list(pdf_dir.glob("*.pdf")) + list(pdf_dir.glob("*.PDF"))

    if not pdf_files:
        print(f"No PDF files found in {pdf_dir}")
        sys.exit(1)

    print(f"Found {len(pdf_files)} PDF files")

    output_dir = Path(output_dir) if output_dir else pdf_dir / "output"

    for pdf_path in pdf_files:
        try:
            extract_single(str(pdf_path), str(output_dir), config_path)
        except Exception as e:
            print(f"Error processing {pdf_path}: {e}")
            continue


def main():
    parser = argparse.ArgumentParser(description="10-K Evidence Extraction")
    subparsers = parser.add_subparsers(dest="command", help="Commands")

    # Extract single file
    extract_parser = subparsers.add_parser("extract", help="Extract from single PDF")
    extract_parser.add_argument("pdf_file", help="Path to 10-K PDF")
    extract_parser.add_argument("--output", "-o", help="Output directory")
    extract_parser.add_argument("--config", "-c", help="Path to custom config JSON file")

    # Batch processing
    batch_parser = subparsers.add_parser("batch", help="Extract from all PDFs in directory")
    batch_parser.add_argument("pdf_dir", help="Directory containing PDFs")
    batch_parser.add_argument("--output", "-o", help="Output directory")
    batch_parser.add_argument("--config", "-c", help="Path to custom config JSON file")

    args = parser.parse_args()

    if args.command == "extract":
        extract_single(args.pdf_file, args.output, args.config)
    elif args.command == "batch":
        extract_batch(args.pdf_dir, args.output, args.config)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
