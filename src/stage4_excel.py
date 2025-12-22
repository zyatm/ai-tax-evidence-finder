"""
Stage 4: Excel Output Builder
=============================
Builds formatted Excel output from extraction and summary results.
Follows SME template specifications.
"""

from pathlib import Path
from datetime import datetime
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

try:
    from .stage2_verbatim import ExtractionResult
    from .stage3_summary import SummaryResult
except ImportError:
    from stage2_verbatim import ExtractionResult
    from stage3_summary import SummaryResult


# Output columns matching SME template
OUTPUT_COLUMNS = [
    ("Potential Method Change List", 45),
    ("found in 10K", 100),
    ("SUMMARIZE", 60),
    ("found in trial balance", 25),
    ("found in tax return", 25),
    ("Potential Benefit or Exposure", 25),
    ("Low", 8),
    ("Mid", 8),
    ("High", 8),
    ("Timing", 12),
    ("Permanent", 12),
]

# Block order for output
BLOCK_ORDER = [
    "Fixed Assets",
    "Inventory",
    "R&D",
    "Tax",
    "Financial Statements"
]


class ExcelBuilder:
    """
    Stage 4: Build formatted Excel output.
    """
    
    def __init__(self, output_path: str):
        self.output_path = Path(output_path)
        self.wb = Workbook()
        
        # Styles
        self.header_font = Font(name='Calibri', size=11, bold=True)
        self.header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        self.block_font = Font(name='Calibri', size=11, bold=True, color='1F4E79')
        self.block_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
        self.normal_font = Font(name='Calibri', size=11)
        self.cell_alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def build(
        self, 
        extraction_result: ExtractionResult,
        summary_result: Optional[SummaryResult] = None,
        source_filename: str = None
    ) -> str:
        """
        Build Excel workbook from extraction results.
        """
        print(f"\n[Stage 4] Building Excel output: {self.output_path}")
        
        # Use first sheet
        ws = self.wb.active
        sheet_name = Path(source_filename or extraction_result.filename).stem[:31]
        ws.title = sheet_name
        
        # Set up headers
        self._write_headers(ws)
        
        # Group categories by block
        categories_by_block = self._group_by_block(extraction_result)
        
        # Write data rows
        row = 2
        for block_name in BLOCK_ORDER:
            if block_name not in categories_by_block:
                continue
            
            # Write block header
            row = self._write_block_header(ws, row, block_name)
            
            # Write category rows
            for cat_id, cat in categories_by_block[block_name].items():
                summary = None
                if summary_result and cat_id in summary_result.summaries:
                    summary = summary_result.summaries[cat_id]
                
                row = self._write_category_row(ws, row, cat, summary)
            
            row += 1  # Empty row between blocks
        
        # Add metadata sheet
        self._add_metadata_sheet(extraction_result, summary_result)
        
        # Save
        self.wb.save(self.output_path)
        print(f"[Stage 4] Saved: {self.output_path}")
        
        return str(self.output_path)
    
    def _write_headers(self, ws):
        """Write column headers."""
        for col, (header, width) in enumerate(OUTPUT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.thin_border
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _write_block_header(self, ws, row: int, block_name: str) -> int:
        """Write block header row."""
        cell = ws.cell(row=row, column=1, value=f"═══ {block_name.upper()} ═══")
        cell.font = self.block_font
        cell.fill = self.block_fill
        
        # Merge across first 3 columns for visibility
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        
        return row + 1
    
    def _write_category_row(self, ws, row: int, cat, summary=None) -> int:
        """Write a category data row."""
        # Column A: Category name
        cell_a = ws.cell(row=row, column=1, value=cat.category_name)
        cell_a.font = self.normal_font
        cell_a.alignment = self.cell_alignment
        
        # Column B: Verbatim evidence
        evidence_text = self._format_evidence(cat.evidence)
        cell_b = ws.cell(row=row, column=2, value=evidence_text)
        cell_b.font = self.normal_font
        cell_b.alignment = self.cell_alignment
        
        # Column C: Summary (if available)
        if summary:
            summary_text = summary.summary
            if summary.tax_opportunities:
                summary_text += f"\n\nOpportunities: {', '.join(summary.tax_opportunities)}"
            if summary.review_flags:
                summary_text += f"\n\n[REVIEW: {', '.join(summary.review_flags)}]"
            
            cell_c = ws.cell(row=row, column=3, value=summary_text)
            cell_c.font = self.normal_font
            cell_c.alignment = self.cell_alignment
        
        # Set row height based on content
        max_lines = max(
            evidence_text.count('\n') + 1 if evidence_text else 1,
            3  # Minimum height
        )
        ws.row_dimensions[row].height = min(max_lines * 15, 400)
        
        return row + 1
    
    def _format_evidence(self, evidence_list) -> str:
        """Format evidence list for Excel cell."""
        if not evidence_list:
            return ""
        
        formatted = []
        for i, e in enumerate(evidence_list, 1):
            # Build evidence entry
            page_info = f"[Page {e.page}]" if e.page else ""
            section_info = f"({e.section})" if e.section else ""
            flags = f" [{', '.join(e.flags)}]" if e.flags else ""
            confidence = f" [Confidence: {e.confidence}]" if e.confidence != "HIGH" else ""
            
            header = f"Evidence {i} {page_info} {section_info}{flags}{confidence}".strip()
            
            # Truncate very long evidence
            text = e.text
            if len(text) > 2000:
                text = text[:2000] + "... [TRUNCATED]"
            
            formatted.append(f"{header}\n{text}")
        
        return "\n\n---\n\n".join(formatted)
    
    def _group_by_block(self, extraction_result: ExtractionResult) -> dict:
        """Group categories by block."""
        blocks = {}
        for cat_id, cat in extraction_result.categories.items():
            block = cat.block
            if block not in blocks:
                blocks[block] = {}
            blocks[block][cat_id] = cat
        return blocks
    
    def _add_metadata_sheet(
        self, 
        extraction_result: ExtractionResult,
        summary_result: Optional[SummaryResult]
    ):
        """Add metadata sheet with extraction info."""
        ws = self.wb.create_sheet(title="Extraction Info")
        
        metadata = [
            ("10-K Extraction Report", ""),
            ("", ""),
            ("Source File", extraction_result.filename),
            ("Extraction Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("", ""),
            ("Extraction Statistics", ""),
            ("Total Evidence Items", extraction_result.total_evidence),
            ("Categories Processed", len(extraction_result.categories)),
            ("Extraction Cost", f"${extraction_result.cost_estimate:.4f}"),
            ("Input Tokens", extraction_result.tokens_used.get('input', 0)),
            ("Output Tokens", extraction_result.tokens_used.get('output', 0)),
        ]
        
        if summary_result:
            metadata.extend([
                ("", ""),
                ("Summary Statistics", ""),
                ("Summaries Generated", len([s for s in summary_result.summaries.values() if "No evidence" not in s.summary])),
                ("Summary Cost", f"${summary_result.cost_estimate:.4f}"),
                ("", ""),
                ("Total Cost", f"${extraction_result.cost_estimate + summary_result.cost_estimate:.4f}"),
            ])
        
        for row, (label, value) in enumerate(metadata, 1):
            ws.cell(row=row, column=1, value=label).font = Font(bold=True if not value else False)
            ws.cell(row=row, column=2, value=value)
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40


def build_excel(
    extraction_result: ExtractionResult,
    output_path: str,
    summary_result: Optional[SummaryResult] = None,
    source_filename: str = None
) -> str:
    """Convenience function to build Excel output."""
    builder = ExcelBuilder(output_path)
    return builder.build(extraction_result, summary_result, source_filename)


def build_multi_excel(
    results: list[tuple[ExtractionResult, Optional[SummaryResult]]],
    output_path: str
) -> str:
    """Build Excel with multiple 10-Ks as separate sheets."""
    builder = ExcelBuilder(output_path)
    
    # Remove default sheet
    builder.wb.remove(builder.wb.active)
    
    for extraction_result, summary_result in results:
        # Create new sheet
        sheet_name = Path(extraction_result.filename).stem[:31]
        ws = builder.wb.create_sheet(title=sheet_name)
        
        # Write headers
        builder._write_headers(ws)
        
        # Group and write data
        categories_by_block = builder._group_by_block(extraction_result)
        row = 2
        
        for block_name in BLOCK_ORDER:
            if block_name not in categories_by_block:
                continue
            
            row = builder._write_block_header(ws, row, block_name)
            
            for cat_id, cat in categories_by_block[block_name].items():
                summary = None
                if summary_result and cat_id in summary_result.summaries:
                    summary = summary_result.summaries[cat_id]
                row = builder._write_category_row(ws, row, cat, summary)
            
            row += 1
    
    # Add combined metadata
    builder._add_metadata_sheet(results[0][0], results[0][1] if results[0][1] else None)
    
    builder.wb.save(output_path)
    return output_path


if __name__ == "__main__":
    print("Stage 4 requires Stage 2/3 output. Run pipeline.py instead.")
